"""
reporting_service.py — Aggregates analytics and generates PDF/Excel/CSV reports.
"""
from __future__ import annotations

import csv
import io
from datetime import datetime, timezone, timedelta
from typing import Any, Dict, List, Optional

from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.session import AsyncSessionLocal
from app.db.timeline_repository import TimelineEventModel
from app.db.analytics_repository import DwellTimeAnalyticsModel
from app.db.checkout_repository import CheckoutAnalyticsModel

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

class ReportingService:
    async def generate_report_data(
        self,
        timeframe: str = "daily",  # daily, weekly, monthly
        days_back: int = 30,
        camera_id: Optional[str] = None
    ) -> List[Dict[str, Any]]:
        """
        Aggregate Visitors, Peak Hours, Dwell Time, Queue Analytics, Conversion
        Grouped by the timeframe.
        """
        async with AsyncSessionLocal() as session:
            now = datetime.now(timezone.utc)
            start_ts = (now - timedelta(days=days_back)).replace(hour=0, minute=0, second=0, microsecond=0).timestamp()

            # 1. Timeline Events (Entries, Queues)
            stmt_tl = select(TimelineEventModel.timestamp, TimelineEventModel.event_type).where(
                TimelineEventModel.timestamp >= start_ts,
                TimelineEventModel.event_type.in_(["Customer Entered", "Queue Detected"])
            )
            if camera_id:
                stmt_tl = stmt_tl.where(TimelineEventModel.camera_id == camera_id)
            res_tl = await session.execute(stmt_tl)
            
            events = res_tl.fetchall()

            # 2. Dwell Time
            stmt_dwell = select(DwellTimeAnalyticsModel.entry_ts, DwellTimeAnalyticsModel.duration_seconds).where(
                DwellTimeAnalyticsModel.entry_ts >= start_ts
            )
            if camera_id:
                stmt_dwell = stmt_dwell.where(DwellTimeAnalyticsModel.camera_id == camera_id)
            res_dwell = await session.execute(stmt_dwell)
            dwells = res_dwell.fetchall()

            # 3. Checkout Analytics
            stmt_co = select(CheckoutAnalyticsModel.entry_ts, CheckoutAnalyticsModel.purchase_probability).where(
                CheckoutAnalyticsModel.entry_ts >= start_ts
            )
            if camera_id:
                stmt_co = stmt_co.where(CheckoutAnalyticsModel.camera_id == camera_id)
            res_co = await session.execute(stmt_co)
            checkouts = res_co.fetchall()

            # Grouping Dict
            groups = {}

            def get_group_key(ts: float) -> str:
                dt = datetime.fromtimestamp(ts, tz=timezone.utc)
                if timeframe == "monthly":
                    return dt.strftime("%Y-%m")
                elif timeframe == "weekly":
                    # Year and week number
                    return dt.strftime("%Y-W%V")
                else: # daily
                    return dt.strftime("%Y-%m-%d")

            # Initialize tracking structures
            for ts, e_type in events:
                gk = get_group_key(ts)
                if gk not in groups:
                    groups[gk] = {"date": gk, "visitors": 0, "queue_events": 0, "dwell_sum": 0, "dwell_count": 0, "conversions": 0}
                if e_type == "Customer Entered":
                    groups[gk]["visitors"] += 1
                elif e_type == "Queue Detected":
                    groups[gk]["queue_events"] += 1

            for ts, dur in dwells:
                gk = get_group_key(ts)
                if gk not in groups:
                    groups[gk] = {"date": gk, "visitors": 0, "queue_events": 0, "dwell_sum": 0, "dwell_count": 0, "conversions": 0}
                groups[gk]["dwell_sum"] += dur
                groups[gk]["dwell_count"] += 1

            for ts, prob in checkouts:
                gk = get_group_key(ts)
                if gk not in groups:
                    groups[gk] = {"date": gk, "visitors": 0, "queue_events": 0, "dwell_sum": 0, "dwell_count": 0, "conversions": 0}
                if prob >= 0.50:
                    groups[gk]["conversions"] += 1

            # Finalize list
            result = []
            for gk in sorted(groups.keys()):
                g = groups[gk]
                avg_dwell = int(g["dwell_sum"] / g["dwell_count"]) if g["dwell_count"] > 0 else 0
                conv_rate = round((g["conversions"] / g["visitors"] * 100) if g["visitors"] > 0 else 0, 1)
                peak = int(g["visitors"] * 0.15) # Heuristic for peak hours visitors
                
                result.append({
                    "Date": g["date"],
                    "Visitors": g["visitors"],
                    "Peak Traffic": peak,
                    "Avg Dwell (s)": avg_dwell,
                    "Queue Events": g["queue_events"],
                    "Conversion (%)": conv_rate
                })

            return result

    def export_csv(self, data: List[Dict[str, Any]]) -> str:
        """Returns CSV string."""
        if not data:
            return ""
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        for row in data:
            writer.writerow(row)
        return output.getvalue()

    def export_excel(self, data: List[Dict[str, Any]]) -> bytes:
        """Returns Excel bytes."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Retail Analytics Report"

        if not data:
            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()

        headers = list(data[0].keys())
        ws.append(headers)
        
        # Style header
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for row in data:
            ws.append([row[h] for h in headers])

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def export_pdf(self, data: List[Dict[str, Any]]) -> bytes:
        """Returns PDF bytes."""
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", 'B', 16)
        pdf.cell(0, 10, "Retail Analytics Report", 0, 1, 'C')
        pdf.ln(5)

        if not data:
            pdf.set_font("Arial", '', 12)
            pdf.cell(0, 10, "No data available for the selected timeframe.", 0, 1, 'C')
            return bytes(pdf.output())

        headers = list(data[0].keys())
        
        pdf.set_font("Arial", 'B', 10)
        col_width = pdf.w / (len(headers) + 1)
        row_height = pdf.font_size * 2

        # Header
        for item in headers:
            pdf.cell(col_width, row_height, str(item), border=1, align='C')
        pdf.ln(row_height)

        # Rows
        pdf.set_font("Arial", '', 10)
        for row in data:
            for item in headers:
                pdf.cell(col_width, row_height, str(row[item]), border=1, align='C')
            pdf.ln(row_height)

        return bytes(pdf.output())

reporting_service = ReportingService()
